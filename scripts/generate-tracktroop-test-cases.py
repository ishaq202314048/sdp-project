#!/usr/bin/env python3
"""Generate TrackTroop test cases Excel in the same format as HallBridge_Test_Cases.xlsx.

- Reads the HallBridge template to preserve sheet name, intro rows, headers, and basic layout.
- Replaces project/module metadata.
- Populates test cases for TrackTroop (auth + role dashboards + key APIs).

This is intentionally dependency-light: uses pandas + openpyxl.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Dict, List

import pandas as pd

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@dataclass(frozen=True)
class TestCase:
    scenario: str
    tc_id: int
    description: str
    test_data: str
    steps: List[str]
    expected: str
    actual: str = ""
    status: str = ""
    created_qa: str = ""
    executed_qa: str = ""


TEMPLATE_PATH = Path("/home/ishaq/Downloads/HallBridge_Test_Cases.xlsx")
OUTPUT_PATH = Path("/home/ishaq/Desktop/tracktroopMergedAuth/tracktroopMergedAuth/TrackTroop_Test_Cases.xlsx")
REPO_ROOT = Path("/home/ishaq/Desktop/tracktroopMergedAuth/tracktroopMergedAuth")


def _find_header_row(df: pd.DataFrame) -> int:
    for i in range(min(100, len(df))):
        row = df.iloc[i].astype(str).tolist()
        if any("Test Scenario" in c for c in row) and any("Test Case ID" in c for c in row):
            return i
    raise RuntimeError("Could not find header row with 'Test Scenario' and 'Test Case ID'.")


def _discover_page_routes(repo_root: Path) -> List[str]:
    """Return sorted unique Next.js page routes based on app/**/page.tsx."""

    def to_route(p: Path) -> str:
        parts = list(p.parts)
        idx = parts.index("app")
        rel = parts[idx + 1 :]
        rel = rel[:-1]  # drop page.tsx
        if not rel:
            return "/"
        return "/" + "/".join(rel)

    routes = {to_route(p) for p in repo_root.glob("app/**/page.tsx")}
    return sorted(routes)


def _discover_api_routes(repo_root: Path) -> List[str]:
    """Return sorted unique API routes based on app/api/**/route.ts."""

    def to_route(p: Path) -> str:
        parts = list(p.parts)
        idx = parts.index("api")
        rel = parts[idx + 1 :]
        rel = rel[:-1]  # drop route.ts
        return "/api/" + "/".join(rel)

    routes = {to_route(p) for p in repo_root.glob("app/api/**/route.ts")}
    return sorted(routes)


def _infer_role_for_page(route: str) -> str:
    if route.startswith("/dashboard/soldier"):
        return "Soldier"
    if route.startswith("/dashboard/clerk"):
        return "Clerk"
    if route.startswith("/dashboard/adjutant"):
        return "Adjutant"
    if route.startswith("/auth"):
        return "Public/Auth"
    return "Public"


def _human_page_name(route: str) -> str:
    if route == "/":
        return "Landing Page"
    if route == "/poster":
        return "Poster Page"
    return route


def _api_scenario_name(api_route: str) -> str:
    # Keep it short but consistent
    if api_route.startswith("/api/auth/"):
        return "Verify Auth API"
    if api_route.startswith("/api/soldiers"):
        return "Verify Soldiers API"
    if api_route.startswith("/api/fitness-test"):
        return "Verify Fitness Test API"
    if api_route.startswith("/api/fitness"):
        return "Verify Fitness API"
    if api_route.startswith("/api/reports"):
        return "Verify Reports API"
    return "Verify API"


def build_test_cases() -> List[TestCase]:
    """Full-coverage TrackTroop test cases aligned to pages + API surface.

    Coverage strategy:
    - Keep the earlier curated end-to-end flows (auth + role dashboards)
    - Add one UI navigation/smoke test per Next.js page
    - Add one happy-path and one auth/role negative test per API route (lightweight, template-friendly)
    """

    cases: List[TestCase] = []
    i = 1

    def add(
        scenario: str,
        description: str,
        test_data: str,
        steps: List[str],
        expected: str,
    ) -> None:
        nonlocal i
        cases.append(
            TestCase(
                scenario=scenario,
                tc_id=i,
                description=description,
                test_data=test_data,
                steps=steps,
                expected=expected,
            )
        )
        i += 1

    # Auth
    add(
        "Verify Login",
        "Login with valid credentials routes user to correct dashboard by role",
        "User: existing (role=Soldier/Clerk/Adjutant)\nPassword: valid",
        [
            "1. Navigate to /auth/login",
            "2. Enter valid username/email and password",
            "3. Click Login",
        ],
        "JWT token is stored (localStorage/session). User is redirected to /dashboard/<role>/home and can access protected pages.",
    )

    add(
        "Verify Login",
        "Login with wrong password shows error and does not create a session",
        "User: existing\nPassword: invalid",
        [
            "1. Navigate to /auth/login",
            "2. Enter valid username/email and an invalid password",
            "3. Click Login",
        ],
        "User stays on login page. Error message appears. No token is stored.",
    )

    add(
        "Verify Signup",
        "Signup with valid soldier data creates account in pending/approval state",
        "Role: Soldier\nPassword: matches policy\nUnique email/username",
        [
            "1. Navigate to /auth/signup",
            "2. Fill in required soldier fields",
            "3. Submit signup",
        ],
        "Account is created. UI indicates pending approval (or redirects to login/pending page). Soldier cannot access dashboards until approved.",
    )

    add(
        "Verify Signup",
        "Signup rejects weak password / invalid input",
        "Role: Soldier\nPassword: weak (missing required characters) OR missing required fields",
        [
            "1. Navigate to /auth/signup",
            "2. Enter invalid data",
            "3. Submit signup",
        ],
        "Inline validation messages show. Request is rejected; no user is created.",
    )

    add(
        "Verify Forgot Password",
        "Forgot password flow accepts email and proceeds through OTP step (simulated)",
        "Email: registered user",
        [
            "1. Navigate to /auth/forgot_password",
            "2. Enter registered email",
            "3. Continue to OTP step",
            "4. Enter OTP (per UI instructions)",
        ],
        "Flow progresses to success state. UI confirms password reset instructions (or simulated success).",
    )

    # Soldier
    add(
        "Verify Soldier Home",
        "Soldier home loads profile summary and tomorrow exercise plan",
        "Role: Soldier\nApproved account\nHas assigned plan",
        [
            "1. Login as approved Soldier",
            "2. Navigate to /dashboard/soldier/home",
        ],
        "Page loads without error. Soldier summary appears. Tomorrow exercises are listed (from /api/soldier-exercises/tomorrow).",
    )

    add(
        "Verify Soldier Fitness",
        "Soldier can view assigned fitness plan and download plan PDF",
        "Role: Soldier\nAssigned plan exists",
        [
            "1. Login as Soldier",
            "2. Navigate to /dashboard/soldier/fitness",
            "3. Click Download/Print plan (if available)",
        ],
        "Assigned plan details render. PDF is generated/downloaded successfully.",
    )

    add(
        "Verify Soldier Progress",
        "Soldier progress page shows fitness test history and trend charts",
        "Role: Soldier\nSome fitness tests exist",
        [
            "1. Login as Soldier",
            "2. Navigate to /dashboard/soldier/progress",
        ],
        "Fitness test history loads (from /api/soldier-fitness-tests). Charts render without overlap/errors.",
    )

    add(
        "Verify Soldier Profile",
        "Soldier profile page displays personal data and computed fitness status",
        "Role: Soldier",
        [
            "1. Login as Soldier",
            "2. Navigate to /dashboard/soldier/profile",
        ],
        "Profile fields show. Fitness status badge/summary renders consistently with stored test results.",
    )

    # Clerk
    add(
        "Verify Clerk Overview",
        "Clerk overview dashboard loads unit overview and new soldier summary",
        "Role: Clerk",
        [
            "1. Login as Clerk",
            "2. Navigate to /dashboard/clerk/overview",
        ],
        "Unit overview loads (from /api/soldiers/unit-overview). Any new-soldier metrics load without errors.",
    )

    add(
        "Verify Clerk Reports",
        "Clerk can generate and send a report (PDF/export) via report endpoint",
        "Role: Clerk\nReport date range: valid",
        [
            "1. Login as Clerk",
            "2. Navigate to /dashboard/clerk/reports",
            "3. Select report range/options",
            "4. Generate report",
        ],
        "Report is generated successfully and can be downloaded/viewed. Server stores report and PDF endpoint works (/api/reports/[id]/pdf).",
    )

    add(
        "Verify Clerk Profile",
        "Clerk profile page loads user info and allows logout",
        "Role: Clerk",
        [
            "1. Login as Clerk",
            "2. Navigate to /dashboard/clerk/profile",
            "3. Click Logout",
        ],
        "Profile loads. Logout clears token/session and redirects to /auth/login.",
    )

    # Adjutant
    add(
        "Verify Adjutant Soldiers",
        "Adjutant can view new/pending soldiers and approve a soldier",
        "Role: Adjutant\nAt least one pending soldier",
        [
            "1. Login as Adjutant",
            "2. Navigate to /dashboard/adjutant/soldiers",
            "3. Select a pending soldier",
            "4. Click Approve",
        ],
        "Approval succeeds. Soldier status updates. Subsequent login as that soldier allows access to dashboards. Backend call /api/soldiers/approve succeeds.",
    )

    add(
        "Verify Adjutant Fitness Tests",
        "Adjutant can create a new IPFT/fitness test record for a soldier",
        "Role: Adjutant\nSoldier: approved\nScores: valid numeric",
        [
            "1. Login as Adjutant",
            "2. Navigate to /dashboard/adjutant/fitness-tests",
            "3. Choose soldier and enter test values",
            "4. Submit",
        ],
        "Test record is created (/api/fitness-test/create). Soldier fitness status recalculates and appears on dashboards.",
    )

    add(
        "Verify Adjutant IPFT Date",
        "Adjutant can set the upcoming IPFT date",
        "Role: Adjutant\nDate: future",
        [
            "1. Login as Adjutant",
            "2. Navigate to /dashboard/adjutant/home",
            "3. Set IPFT date",
            "4. Save",
        ],
        "Saved date persists (/api/fitness-test/ipft-date) and is displayed to other roles if surfaced.",
    )

    add(
        "Verify Weekly Plan Assignment",
        "Adjutant assigns a fitness plan to a soldier; soldier can see it",
        "Role: Adjutant\nPlan: existing\nSoldier: approved",
        [
            "1. Login as Adjutant",
            "2. Navigate to /dashboard/adjutant/home (or plan screen)",
            "3. Select plan and soldier",
            "4. Assign plan",
            "5. Login as that Soldier and open /dashboard/soldier/fitness",
        ],
        "Assignment succeeds (/api/fitness/assign). Soldier sees the assigned plan immediately.",
    )

    add(
        "Verify High-Risk Alerts",
        "Adjutant alert page lists high-risk soldiers based on thresholds",
        "Role: Adjutant\nAt least one soldier low score / high risk",
        [
            "1. Login as Adjutant",
            "2. Navigate to /dashboard/adjutant/alert",
        ],
        "High-risk list loads from /api/high-risk-soldiers. Data looks consistent with latest fitness tests.",
    )

    add(
        "Verify Reports & Analytics",
        "Adjutant reports dashboard loads charts and monthly trends",
        "Role: Adjutant\nHas historical data",
        [
            "1. Login as Adjutant",
            "2. Navigate to /dashboard/adjutant/reports",
        ],
        "Charts render from /api/fitness/reports-chart and /api/fitness/monthly-trend with correct axes and no client crashes.",
    )

    # API-focused negative/security tests (kept aligned to typical template)
    add(
        "Verify Authorization",
        "Unauthenticated user cannot access protected API endpoints",
        "No Authorization header / no token",
        [
            "1. Call /api/profile without auth",
            "2. Call /api/soldiers without auth",
        ],
        "API responds 401/403. No sensitive data is returned.",
    )

    add(
        "Verify Authorization",
        "Role-based restriction: Soldier cannot approve soldiers",
        "Role: Soldier (valid token)",
        [
            "1. Login as Soldier",
            "2. Attempt POST /api/soldiers/approve",
        ],
        "API responds 403 (or equivalent). No approval occurs.",
    )

    # ---------------------------------------------------------------------
    # Full coverage: Page smoke tests (one per route)
    # ---------------------------------------------------------------------
    page_routes = _discover_page_routes(REPO_ROOT)
    for route in page_routes:
        role = _infer_role_for_page(route)
        name = _human_page_name(route)
        scenario = "Verify Frontend Page"
        desc = f"Open {name} and verify it loads and shows expected UI sections"
        test_data = f"User role: {role}"
        steps = [
            f"1. Navigate to {route}",
            "2. Observe page load and any redirects",
            "3. Confirm primary widgets/sections render without runtime errors",
        ]
        expected = (
            "Page loads successfully (or redirects appropriately). "
            "No blank screen or console crash. Core content/components are visible."
        )
        add(scenario, desc, test_data, steps, expected)

    # ---------------------------------------------------------------------
    # Full coverage: API endpoint tests (happy path + auth negative)
    # ---------------------------------------------------------------------
    api_routes = _discover_api_routes(REPO_ROOT)
    # minimal role mapping for negative tests
    role_for_api: Dict[str, str] = {
        "/api/soldiers/approve": "Soldier",
    }

    for api in api_routes:
        scenario = _api_scenario_name(api)

        # Happy-path/general call
        add(
            scenario,
            f"Successful request to {api} returns expected response shape",
            "Auth: valid token (as required by endpoint)\nQuery/body: valid",
            [
                "1. Ensure you are authenticated with appropriate role",
                f"2. Send request to {api} with valid inputs",
                "3. Inspect HTTP status and JSON/PDF response",
            ],
            "Response status is 200/201 (as applicable). Body matches expected schema (or PDF downloads for pdf endpoints).",
        )

        # Unauthenticated/forbidden check
        negative_role = role_for_api.get(api, "Unauthenticated")
        add(
            scenario,
            f"Authorization enforced for {api} (no token / wrong role)",
            f"Auth: {negative_role}",
            [
                f"1. Call {api} without token (or using wrong role token)",
                "2. Observe HTTP status and error body",
            ],
            "API responds 401/403 (as applicable). No sensitive data is returned and no state-changing action is performed.",
        )

    return cases


def main() -> None:
    if not TEMPLATE_PATH.exists():
        raise SystemExit(f"Template not found: {TEMPLATE_PATH}")

    # Load raw template (keep layout as much as possible)
    raw = pd.read_excel(TEMPLATE_PATH, sheet_name=0, header=None)
    header_row = _find_header_row(raw)
    sheet_name = pd.ExcelFile(TEMPLATE_PATH).sheet_names[0]

    # Update metadata cells (based on observed template layout)
    # Row 1 col 1: Project Name label is in col 0, value is in col 1
    # We keep labels and replace values.
    def set_value(row_idx: int, value: str) -> None:
        if raw.shape[1] < 2:
            return
        raw.iat[row_idx, 1] = value

    set_value(1, "TrackTroop - Troop Fitness Tracking System")
    set_value(
        2,
        "Auth, Soldier Dashboard, Clerk Dashboard, Adjutant Dashboard, Fitness Plans, Fitness Tests, Reports",
    )
    set_value(3, "Generated (AI-assisted)")
    set_value(4, date.today().strftime("%B %d, %Y"))

    # Build and inject test cases
    cases = build_test_cases()

    # Ensure we have enough rows to write all cases (template may be shorter than our expanded catalog)
    # Worst case per test case: 1 header row + (1 base row + (len(steps)-1) step rows + 1 spacer row)
    required_rows = header_row + 1
    for tc in cases:
        required_rows += max(1, len(tc.steps)) + 1
    if required_rows > len(raw):
        extra = required_rows - len(raw)
        raw = pd.concat(
            [raw, pd.DataFrame([""] * raw.shape[1] for _ in range(extra))],
            ignore_index=True,
        )

    # Clear existing test case rows in the template area (from header_row+1 onwards)
    # but preserve the header row. We'll blank out everything below header_row.
    raw.iloc[header_row + 1 :, :] = ""

    # Template columns in order
    columns = [
        "Test Scenario",
        "Test Case ID",
        "Test Case Description",
        "Test Data",
        "Steps to Execute",
        "Expected Result",
        "Actual Result",
        "Status",
        "Created QA Name",
        "Executed QA Name",
    ]

    # Write header row explicitly (some templates may have merged cells; we still fill)
    for col_idx, name in enumerate(columns):
        raw.iat[header_row, col_idx] = name

    # Fill rows
    r = header_row + 1
    for tc in cases:
        raw.iat[r, 0] = tc.scenario
        raw.iat[r, 1] = tc.tc_id
        raw.iat[r, 2] = tc.description
        raw.iat[r, 3] = tc.test_data
        # Steps: first step on the same row, subsequent steps on following rows
        raw.iat[r, 4] = tc.steps[0] if tc.steps else ""
        raw.iat[r, 5] = tc.expected
        raw.iat[r, 6] = tc.actual
        raw.iat[r, 7] = tc.status
        raw.iat[r, 8] = tc.created_qa
        raw.iat[r, 9] = tc.executed_qa
        r += 1
        for step in tc.steps[1:]:
            raw.iat[r, 4] = step
            r += 1
        # add a blank spacer row like the template often does
        r += 1

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    # Write with openpyxl engine
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        raw.to_excel(writer, sheet_name=sheet_name, header=False, index=False)

    # ------------------------------------------------------------------
    # Formatting pass (make it presentable like a real QA sheet)
    # ------------------------------------------------------------------
    wb = load_workbook(OUTPUT_PATH)
    ws = wb[sheet_name]

    # Global font (Calibri-like is standard for Excel)
    base_font = Font(name="Calibri", size=11)
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    # Column widths (copied from the HallBridge template)
    col_widths = {
        "A": 22.0,
        "B": 14.0,
        "C": 35.0,
        "D": 30.0,
        "E": 35.0,
        "F": 30.0,
        "G": 25.0,
        "H": 12.0,
        "I": 18.0,
        "J": 18.0,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    thin = Side(style="thin", color="9CA3AF")  # soft gray
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    body_alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    id_alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

    header_fill = PatternFill("solid", fgColor="1F4E79")  # professional blue
    zebra_fill = PatternFill("solid", fgColor="F3F6FA")  # very light gray-blue

    # header_row is 0-based in pandas; Excel is 1-based
    header_row_excel = header_row + 1

    # Style header
    for c in range(1, 11):
        cell = ws.cell(row=header_row_excel, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Add auto-filter on header row
    ws.auto_filter.ref = f"A{header_row_excel}:J{header_row_excel}"

    # Style body rows (from first data row to last used row)
    last_row = ws.max_row
    for r in range(header_row_excel + 1, last_row + 1):
        # Zebra striping on main rows that have a test case id
        is_main_row = ws.cell(row=r, column=2).value not in (None, "")
        if is_main_row and (int(ws.cell(row=r, column=2).value) % 2 == 0):
            row_fill = zebra_fill
        else:
            row_fill = None

        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            # keep metadata area untouched (above header row)
            cell.border = border
            cell.font = base_font
            if row_fill is not None:
                cell.fill = row_fill
            if c == 2:  # Test Case ID
                cell.alignment = id_alignment
            elif c in (1, 8):
                # Scenario + Status looks nicer centered vertically
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
            else:
                cell.alignment = body_alignment

        # auto-ish row height: keep readable for wrapped text
        ws.row_dimensions[r].height = 45

    # Keep top metadata compact
    for r in range(1, header_row_excel):
        ws.row_dimensions[r].height = 18

    # Freeze the header row (so teacher can scroll)
    ws.freeze_panes = f"A{header_row_excel + 1}"

    # Print settings (makes PDF/printouts look clean)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_title_rows = f"{header_row_excel}:{header_row_excel}"
    ws.sheet_view.showGridLines = False

    # Conditional formatting for Status column (H): Passed/Failed/Blocked
    # Apply from first data row to last_row
    status_range = f"$H${header_row_excel+1}:$H${last_row}"
    passed_fill = PatternFill("solid", fgColor="C6EFCE")
    failed_fill = PatternFill("solid", fgColor="FFC7CE")
    blocked_fill = PatternFill("solid", fgColor="FFEB9C")
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(formula=[f'UPPER($H{header_row_excel+1})="PASSED"'], fill=passed_fill),
    )
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(formula=[f'UPPER($H{header_row_excel+1})="FAILED"'], fill=failed_fill),
    )
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(formula=[f'UPPER($H{header_row_excel+1})="BLOCKED"'], fill=blocked_fill),
    )

    wb.save(OUTPUT_PATH)

    print(f"Wrote: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
